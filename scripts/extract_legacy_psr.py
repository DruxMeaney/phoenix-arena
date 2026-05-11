#!/usr/bin/env python3
"""Extract Phoenix legacy Excel tournaments into a normalized PSR import JSON.

The historical workbooks live outside the app repo and are not consistent
enough to import directly into Prisma. This script creates an auditable staging
file that preserves source workbook, teams, players, per-map kills, placement,
legacy points and verification flags.

Usage:
  python3 scripts/extract_legacy_psr.py \
    --source ../00_old \
    --output data/legacy-psr/legacy-import.json
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
from collections import Counter, defaultdict
from datetime import UTC, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


INVALID_HANDLES = {
    "",
    "#error!",
    "0",
    "na",
    "n/a",
    "none",
    "null",
    "tbd",
    "total",
    "lugar",
    "multiplicador",
    "puntos",
    "jugador",
    "player",
}

INVALID_HANDLE_PATTERNS = [
    re.compile(r"^(equipo|team)\s*\d+[._-]\d+$"),
    re.compile(r"^(jugador|player)\s*\d+$"),
    re.compile(r"^(capitan|capitan manual|team\s*[234]\s*manual)$"),
]


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    text = re.sub(r"\s+", " ", text)
    return text


def normalized_handle(value: Any) -> str:
    text = clean_text(value).lower()
    text = re.sub(r"[^a-z0-9_ .-]+", "", text)
    text = re.sub(r"\s+", " ", text).strip()
    return text


def normalized_key(value: Any) -> str:
    text = normalized_handle(value)
    text = text.replace("team ", "")
    return re.sub(r"[^a-z0-9]+", "", text)


def is_valid_handle(value: Any) -> bool:
    norm = normalized_handle(value)
    if norm in INVALID_HANDLES or len(norm) < 2:
        return False
    return not any(pattern.match(norm) for pattern in INVALID_HANDLE_PATTERNS)


def placeholder_slot(value: Any) -> int | None:
    match = re.match(r"^(?:equipo|team)\s*\d+[._-](\d+)$", normalized_handle(value))
    if not match:
        return None
    return int(match.group(1))


def num(value: Any, default: float = 0) -> float:
    if value is None or value == "":
        return default
    try:
        if isinstance(value, str) and value.strip().startswith("#"):
            return default
        parsed = float(value)
        return parsed if parsed == parsed else default
    except (TypeError, ValueError):
        return default


def int_or_none(value: Any) -> int | None:
    parsed = num(value, -1)
    return int(parsed) if parsed >= 0 else None


def parse_event_date(file_name: str, fallback_index: int, default_year: int) -> str:
    match = re.match(r"^(\d{1,2})[-_](\d{1,2})", file_name)
    if match:
        month = int(match.group(1))
        day = int(match.group(2))
        if 1 <= month <= 12 and 1 <= day <= 31:
            return datetime(default_year, month, day, 20, 0).isoformat()
    return datetime(default_year, 1, 1, 20, 0).replace(day=1).isoformat()


def infer_tournament_type(name: str) -> str:
    lower = name.lower()
    if "novice" in lower:
        return "novice"
    if "scrim" in lower or "tryout" in lower or "wsow" in lower or "ewc" in lower:
        return "scrim"
    if "only detri" in lower or "only detry" in lower or "only detris" in lower:
        return "only_detri"
    if "all skill" in lower or "allskills" in lower:
        return "all_skills"
    if "pro" in lower or "am" in lower:
        return "pro_am_detri"
    if "community" in lower:
        return "community"
    if "coed" in lower:
        return "coed"
    return "legacy_custom"


def get_sheet(workbook, *names: str):
    lower_map = {sheet.lower(): sheet for sheet in workbook.sheetnames}
    for name in names:
        actual = lower_map.get(name.lower())
        if actual:
            return workbook[actual]
    return None


def row_values(row) -> list[Any]:
    return [cell for cell in row]


def parse_registro(workbook) -> dict[str, dict[str, Any]]:
    ws = get_sheet(workbook, "Registro")
    if ws is None:
        return {}

    rows = list(ws.iter_rows(values_only=True))
    header_index = None
    headers: list[str] = []
    for idx, row in enumerate(rows[:12]):
        values = [clean_text(value) for value in row]
        if "No Equipo" in values and any("Equipo" in value for value in values):
            header_index = idx
            headers = values
            break
    if header_index is None:
        return {}

    by_number: dict[str, dict[str, Any]] = {}
    for row in rows[header_index + 1 :]:
        values = row_values(row)
        team_number = None
        try:
            team_number = int(num(values[1], -1))
        except IndexError:
            pass
        if team_number is None or team_number < 0:
            continue

        indexed = {headers[i]: values[i] if i < len(values) else None for i in range(len(headers))}
        team_name = clean_text(indexed.get("Equipo Manual") or indexed.get("Equipo") or "")
        captain = clean_text(indexed.get("Capitan Manual") or indexed.get("Capitan") or "")
        team_2 = clean_text(indexed.get("Team 2 Manual") or indexed.get("Team 2") or "")
        team_3 = clean_text(indexed.get("Team 3 Manual") or indexed.get("Team 3") or "")
        team_4 = clean_text(indexed.get("Team 4 Manual") or indexed.get("Team 4") or "")
        group = clean_text(indexed.get("Grupo") or "")
        by_number[str(team_number)] = {
            "teamName": team_name,
            "captainHandle": captain,
            "players": [p for p in [captain, team_2, team_3, team_4] if is_valid_handle(p)],
            "group": group,
            "paymentVerified": bool(indexed.get("Pago")) if "Pago" in indexed else False,
            "discordVerified": bool(indexed.get("Discord")) if "Discord" in indexed else False,
            "photoVerified": bool(indexed.get("Foto")) if "Foto" in indexed else False,
            "flyerVerified": bool(indexed.get("Flyer")) if "Flyer" in indexed else False,
        }
    return by_number


def parse_position_table(workbook) -> dict[str, int]:
    ws = get_sheet(workbook, "Tabla de posiciones")
    if ws is None:
        return {}

    placements: dict[str, int] = {}
    for row in ws.iter_rows(values_only=True):
        values = row_values(row)
        if len(values) < 3:
            continue
        top = int_or_none(values[1])
        team = clean_text(values[2])
        if top is None or top <= 0 or not team or team.lower() == "equipo":
            continue
        placements[normalized_key(team)] = top
    return placements


def map_values(values: list[Any], start_col: int = 3) -> list[float]:
    return [num(value, 0) for value in values[start_col:]]


def parse_resultados(workbook, registro: dict[str, dict[str, Any]], placements: dict[str, int]) -> list[dict[str, Any]]:
    ws = get_sheet(workbook, "Resultados")
    if ws is None:
        return []

    rows = [row_values(row) for row in ws.iter_rows(values_only=True)]
    teams: list[dict[str, Any]] = []
    i = 0
    while i < len(rows):
        row = rows[i]
        team_name = clean_text(row[0] if len(row) > 0 else "")
        label = clean_text(row[2] if len(row) > 2 else "")
        if not team_name or team_name.lower() == "nombre" or not (is_valid_handle(label) or placeholder_slot(label)):
            i += 1
            continue

        team_number = int_or_none(row[1] if len(row) > 1 else None)
        reg = registro.get(str(team_number or ""), {})
        players: list[dict[str, Any]] = []
        totals: list[float] = []
        places: list[float] = []
        multipliers: list[float] = []
        points: list[float] = []
        j = i
        while j < len(rows):
            current = rows[j]
            current_team = clean_text(current[0] if len(current) > 0 else "")
            current_label = clean_text(current[2] if len(current) > 2 else "")
            lower_label = current_label.lower()

            if j > i and current_team and lower_label not in {"total", "lugar", "multiplicador", "puntos"}:
                break

            if lower_label == "total":
                totals = map_values(current)
            elif lower_label == "lugar":
                places = map_values(current)
            elif lower_label == "multiplicador":
                multipliers = map_values(current)
            elif lower_label == "puntos":
                points = map_values(current)
            else:
                candidate_handle = current_label
                slot = placeholder_slot(current_label)
                registered_players = reg.get("players", [])
                if slot and 1 <= slot <= len(registered_players):
                    candidate_handle = registered_players[slot - 1]
                if not is_valid_handle(candidate_handle):
                    j += 1
                    continue
                players.append(
                    {
                        "rawHandle": candidate_handle,
                        "normalizedHandle": normalized_handle(candidate_handle),
                        "killsByMap": map_values(current),
                    }
                )
            j += 1

        if players:
            team_key = normalized_key(team_name)
            final_placement = placements.get(team_key)
            if final_placement is None:
                reg_key = normalized_key(reg.get("teamName", ""))
                final_placement = placements.get(reg_key)

            max_maps = max(
                [len(points), len(places), len(totals), *[len(player["killsByMap"]) for player in players]],
                default=0,
            )
            round_results = []
            for map_index in range(max_maps):
                team_kills = totals[map_index] if map_index < len(totals) else 0
                placement = places[map_index] if map_index < len(places) else 0
                multiplier = multipliers[map_index] if map_index < len(multipliers) else 0
                skill_points = points[map_index] if map_index < len(points) else team_kills * multiplier
                if team_kills or placement or skill_points:
                    round_results.append(
                        {
                            "map": map_index + 1,
                            "teamKills": team_kills,
                            "placement": placement,
                            "placementMultiplier": multiplier,
                            "skillPoints": skill_points,
                        }
                    )

            rounds_played = len(round_results)
            team_kills_total = sum(item["teamKills"] for item in round_results)
            skill_points_total = sum(item["skillPoints"] for item in round_results)
            average_placement = (
                sum(item["placement"] for item in round_results if item["placement"] > 0)
                / max(1, len([item for item in round_results if item["placement"] > 0]))
            )
            teams.append(
                {
                    "teamName": team_name,
                    "teamNumber": team_number,
                    "teamGroup": reg.get("group", ""),
                    "placement": final_placement,
                    "roundsPlayed": rounds_played,
                    "averagePlacement": round(average_placement, 4) if average_placement else final_placement,
                    "teamKills": round(team_kills_total, 4),
                    "skillPoints": round(skill_points_total, 4),
                    "rawPoints": round(skill_points_total, 4),
                    "matchpointWin": False,
                    "paymentVerified": reg.get("paymentVerified", False),
                    "discordVerified": reg.get("discordVerified", False),
                    "photoVerified": reg.get("photoVerified", False),
                    "flyerVerified": reg.get("flyerVerified", False),
                    "roundResults": round_results,
                    "players": players,
                }
            )
        i = max(j, i + 1)

    missing = [team for team in teams if not team.get("placement")]
    if missing:
        ranked = sorted(teams, key=lambda item: item.get("skillPoints", 0), reverse=True)
        for rank, team in enumerate(ranked, 1):
            if not team.get("placement"):
                team["placement"] = rank
    return teams


def summarize_players(events: list[dict[str, Any]]) -> list[dict[str, Any]]:
    by_handle: dict[str, dict[str, Any]] = {}
    aliases: defaultdict[str, Counter] = defaultdict(Counter)
    for event in events:
        for team in event["teams"]:
            for player in team["players"]:
                norm = player["normalizedHandle"]
                if not is_valid_handle(norm):
                    continue
                kills = sum(player.get("killsByMap", []))
                current = by_handle.setdefault(
                    norm,
                    {
                        "normalizedHandle": norm,
                        "displayHandle": player["rawHandle"],
                        "participations": 0,
                        "kills": 0,
                        "skillPoints": 0,
                        "events": 0,
                    },
                )
                aliases[norm][player["rawHandle"]] += 1
                current["participations"] += 1
                current["kills"] += kills
                current["skillPoints"] += team.get("skillPoints", 0)
                current["events"] += 1
    players = []
    for norm, payload in by_handle.items():
        display = aliases[norm].most_common(1)[0][0] if aliases[norm] else payload["displayHandle"]
        players.append({**payload, "displayHandle": display, "aliases": dict(aliases[norm])})
    return sorted(players, key=lambda item: (item["participations"], item["kills"]), reverse=True)


def extract(source: Path, output: Path, default_year: int, max_files: int | None) -> dict[str, Any]:
    files = sorted([p for p in source.glob("*.xlsx") if not p.name.startswith("~$")])
    if max_files:
        files = files[:max_files]

    events: list[dict[str, Any]] = []
    skipped: list[dict[str, str]] = []
    for index, file_path in enumerate(files):
        try:
            workbook = load_workbook(file_path, read_only=True, data_only=True)
            registro = parse_registro(workbook)
            placements = parse_position_table(workbook)
            teams = parse_resultados(workbook, registro, placements)
            if not teams:
                skipped.append({"fileName": file_path.name, "reason": "no importable Resultados blocks"})
                continue

            source_id = hashlib.sha256(file_path.name.encode("utf-8")).hexdigest()[:16]
            total_teams = len(teams)
            map_count = max((team.get("roundsPlayed", 0) for team in teams), default=0)
            events.append(
                {
                    "sourceFile": str(file_path),
                    "fileName": file_path.name,
                    "sourceId": f"legacy-{source_id}",
                    "name": file_path.stem,
                    "occurredAt": parse_event_date(file_path.name, index, default_year),
                    "tournamentType": infer_tournament_type(file_path.name),
                    "totalTeams": total_teams,
                    "mapCount": map_count,
                    "teams": teams,
                }
            )
        except Exception as error:  # noqa: BLE001 - extractor should keep going.
            skipped.append({"fileName": file_path.name, "reason": str(error)})

    players = summarize_players(events)
    stats = {
        "filesScanned": len(files),
        "eventsImported": len(events),
        "filesSkipped": len(skipped),
        "teamsImported": sum(len(event["teams"]) for event in events),
        "playerParticipations": sum(
            len(team["players"]) for event in events for team in event["teams"]
        ),
        "uniquePlayers": len(players),
    }
    payload = {
        "version": "legacy-psr-import-0.1",
        "generatedAt": datetime.now(UTC).isoformat().replace("+00:00", "Z"),
        "sourceRoot": str(source),
        "stats": stats,
        "players": players,
        "events": events,
        "skipped": skipped,
    }
    output.parent.mkdir(parents=True, exist_ok=True)
    output.write_text(json.dumps(payload, ensure_ascii=True, indent=2), encoding="utf-8")
    return payload


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--source", type=Path, default=Path("../00_old"))
    parser.add_argument("--output", type=Path, default=Path("data/legacy-psr/legacy-import.json"))
    parser.add_argument("--year", type=int, default=2025)
    parser.add_argument("--max-files", type=int, default=0)
    args = parser.parse_args()

    payload = extract(args.source, args.output, args.year, args.max_files or None)
    print(json.dumps(payload["stats"], indent=2))
    if payload["skipped"]:
        print(f"Skipped {len(payload['skipped'])} files. See output JSON for details.")


if __name__ == "__main__":
    main()
